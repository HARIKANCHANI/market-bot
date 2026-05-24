#!/usr/bin/env python3
"""Quickly inspect top-ranked stocks with institutional metrics.

Usage:
  python scripts/analyze_institutional_top.py [optional_path_to_excel]

If no path is provided, the script looks for the latest
`market_analysis_*.xlsx` file under `src/bots/excel`.
"""

import os
import sys
import glob

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_DIR = os.path.join(BASE_DIR, "src", "bots", "excel")


def find_latest_full_report() -> str | None:
    pattern = os.path.join(EXCEL_DIR, "market_analysis_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def analyze_report(path: str, top_n: int = 20) -> None:
    if not os.path.exists(path):
        print(f"❌ Report file not found: {path}")
        return

    print(f"📄 Using report: {path}")

    wb = load_workbook(path, data_only=True)
    if "Stock Analysis" not in wb.sheetnames:
        print("❌ Sheet 'Stock Analysis' not found in workbook")
        return

    ws = wb["Stock Analysis"]

    # Header row
    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {name: idx for idx, name in enumerate(header_values) if name}

    columns = [
        "Rank",
        "Symbol",
        "Sector",
        "Score",
        "Institutional Score",
        "Institutional Confidence",
        "FII Holdings %",
        "DII Holdings %",
        "Promoter Holdings %",
        "Mutual Fund Holdings %",
        "FII Trend",
        "DII Trend",
    ]

    missing = [c for c in columns if c not in col_index]
    if missing:
        print("❌ Missing expected columns in Excel header:", ", ".join(missing))
        return

    def fmt(value):
        if value is None:
            return "N/A"
        if isinstance(value, float):
            return f"{value:.1f}"
        return str(value)

    print("\nTop stocks by Rank with institutional metrics:\n")
    print(" | ".join(columns))
    print("-" * 120)

    for row in ws.iter_rows(min_row=2, max_row=top_n + 1, values_only=True):
        rank_val = row[col_index["Rank"]]
        if rank_val is None:
            continue

        out = []
        for col in columns:
            idx = col_index[col]
            out.append(fmt(row[idx]))
        print(" | ".join(out))


def main() -> None:
    if len(sys.argv) > 1:
        report_path = sys.argv[1]
    else:
        report_path = find_latest_full_report()
        if not report_path:
            print(
                "❌ No full market_analysis_*.xlsx report found in "
                f"{EXCEL_DIR}. Run src/bots/market_bot_excel.py first."
            )
            return

    analyze_report(report_path, top_n=20)


if __name__ == "__main__":
    main()
