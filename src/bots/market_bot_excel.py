"""
EXCEL VERSION - Market Bot (No Notion Upload)
Creates comprehensive Excel reports with all stock data including FII/DII information
Saves timestamped Excel files to src/bots/excel/ folder
Uses same features as LITE version: technical indicators (momentum & volume)
Includes intelligent multi-factor ranking system.
"""

import os
import sys
import time
import re
from datetime import datetime
from urllib.parse import quote
import requests
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Add project root to Python path
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Setup centralized logging
from src.config.logging_config import setup_bot_logging
logger = setup_bot_logging("market_bot_excel")

# Import stock data (prefer 650-stock version if available)
try:
    from data.nse_stocks_650 import (
        get_all_stocks_with_classification,
        get_validated_stocks,
        get_current_ticker,
        is_delisted
    )
except ImportError:
    print("⚠️  Stock data module not found. Please check data/nse_stocks_650.py")
    get_validated_stocks = None
    get_all_stocks_with_classification = None
    get_current_ticker = lambda x: x
    is_delisted = lambda x: False

# Import comprehensive news sources
try:
    from src.core.news_aggregator import fetch_comprehensive_news as fetch_news_comprehensive
    HAS_COMPREHENSIVE_NEWS = True
except ImportError:
    HAS_COMPREHENSIVE_NEWS = False
    print("⚠️  Comprehensive news module not found. Using basic news.")

# Import analyst ratings
try:
    from src.core.analyst_ratings import aggregate_all_analyst_ratings
    HAS_ANALYST_RATINGS = True
except ImportError:
    HAS_ANALYST_RATINGS = False
    print("⚠️  Analyst ratings module not found. Ratings will not be included.")

# Import intelligent ranking engine
try:
    from src.core.ranking_engine import rank_stocks
    HAS_RANKING_ENGINE = True
except ImportError:
    HAS_RANKING_ENGINE = False
    print("⚠️  Ranking engine not found. Using serial ranking.")

# News sentiment keywords
POSITIVE_KEYWORDS = [
    "wins", "won", "award", "beats", "surges", "jumps", "rallies", "gains",
    "record", "profit", "growth", "dividend", "expansion", "order", "contract",
    "upgrade", "acquisition", "launch", "partnership"
]

NEGATIVE_KEYWORDS = [
    "falls", "drops", "crashes", "plunges", "loss", "downgrade", "weak",
    "concern", "debt", "penalty", "resignation", "lawsuit", "decline"
]

# News type classification keywords
NEWS_TYPE_KEYWORDS = {
    "Earnings": ["earnings", "quarter", "q1", "q2", "q3", "q4", "revenue", "profit", "loss", "results"],
    "Product": ["launch", "product", "unveils", "introduces", "release", "model", "variant"],
    "Legal": ["lawsuit", "court", "legal", "case", "trial", "settlement", "penalty", "fine"],
    "M&A": ["merger", "acquisition", "acquires", "deal", "buyout", "takeover", "stake"],
    "Management": ["ceo", "cfo", "resign", "appoint", "director", "board", "executive"],
    "Dividend": ["dividend", "payout", "shareholder", "distribution"],
    "Regulatory": ["sebi", "regulator", "compliance", "violation", "probe", "investigation"],
    "Expansion": ["expansion", "plant", "facility", "capacity", "invest", "capex"]
}

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

# Reuse a single HTTP session for all outbound requests (NSE, Google News,
# Screener). This improves performance when generating large Excel reports.
SESSION = requests.Session()

print("✅ Excel Market Bot Started (Technical Analysis + Excel Export)")

def fetch_news(ticker):
    """Fetch news from Yahoo Finance and Google News"""
    try:
        # Use current ticker (handles renamed companies)
        current_ticker = get_current_ticker(ticker)

        stock = yf.Ticker(f"{current_ticker}.NS")
        yf_news = stock.news or []

        # Get company info
        info = stock.info
        company_name = info.get("longName") or info.get("shortName") or ticker

        all_news = []

        # Yahoo Finance news
        for item in yf_news[:8]:
            title = item.get("title", "")
            pub_time = item.get("providerPublishTime", 0)
            date = datetime.fromtimestamp(pub_time).strftime("%d-%b") if pub_time else "Recent"
            all_news.append({"title": title, "date": date})

        # Google News RSS
        try:
            url = f"https://news.google.com/rss/search?q={quote(company_name)}+stock+india&hl=en-IN&gl=IN&ceid=IN:en"
            response = SESSION.get(
                url,
                timeout=5,
                headers={'User-Agent': USER_AGENT},
            )

            if response.status_code == 200:
                titles = re.findall(r'<title><!\[CDATA\[(.*?)\]\]></title>', response.text)
                if not titles:
                    titles = re.findall(r'<title>(.*?)</title>', response.text)

                for title in titles[1:4]:
                    all_news.append({"title": title, "date": "Recent"})
        except Exception:
            # Ignore Google News failures; Yahoo Finance news is primary
            pass

        # Return top 3 news
        if all_news:
            news_items = []
            for news in all_news[:3]:
                title = re.sub(r'\s+', ' ', news["title"]).strip()
                title = re.sub(r'&amp;', '&', title)
                if len(title) > 150:
                    title = title[:147] + "..."
                news_items.append(f"[{news['date']}] {title}")

            return " | ".join(news_items), [n["title"] for n in all_news[:5]]

        return None, []

    except Exception as e:
        print(f"⚠️  Error fetching news for {ticker}: {str(e)}")
        return None, []

def analyze_news_sentiment(news_text):
    """Analyze news sentiment using keywords"""
    if not news_text:
        return "Neutral"

    news_lower = news_text.lower()
    pos_count = sum(1 for word in POSITIVE_KEYWORDS if word in news_lower)
    neg_count = sum(1 for word in NEGATIVE_KEYWORDS if word in news_lower)

    if pos_count > neg_count:
        return "Positive"
    elif neg_count > pos_count:
        return "Negative"
    return "Neutral"

def classify_news_type(news_text):
    """Classify news into types using keywords"""
    if not news_text:
        return []

    news_lower = news_text.lower()
    types = []

    for news_type, keywords in NEWS_TYPE_KEYWORDS.items():
        if any(keyword in news_lower for keyword in keywords):
            types.append(news_type)

    # Return unique types, max 3
    return list(set(types))[:3] if types else []

def fetch_price_from_nse(symbol):
    """Fetch latest price and basic info from NSE's official quote API.

    Returns a dict with keys: price, pchange, volume, sector, market_cap
    or None if the request fails.
    """

    base_url = "https://www.nseindia.com/api/quote-equity"
    params = {"symbol": symbol}
    headers = {
        "User-Agent": USER_AGENT,
        "Referer": f"https://www.nseindia.com/get-quotes/equity?symbol={quote(symbol)}",
        "Accept": "application/json, text/plain, */*",
    }

    try:
        resp = SESSION.get(base_url, params=params, headers=headers, timeout=8)
        if resp.status_code != 200:
            print(f"   ⚠️ NSE quote HTTP {resp.status_code} for {symbol}")
            return None

        data = resp.json()
        price_info = data.get("priceInfo") or {}
        last_price = price_info.get("lastPrice")
        if last_price is None:
            return None

        pchange = price_info.get("pChange")

        pre_open = data.get("preOpenMarket") or {}
        volume = None
        if isinstance(pre_open, dict):
            volume = pre_open.get("totalTradedVolume")
        if volume is None:
            volume = price_info.get("totalTradedVolume")

        industry_info = data.get("industryInfo") or {}
        info = data.get("info") or {}
        sector = industry_info.get("sector") or info.get("industry") or "Unknown"

        sec_info = data.get("securityInfo") or {}
        issued_size = sec_info.get("issuedSize")
        market_cap = None
        try:
            if issued_size is not None:
                market_cap_raw = float(last_price) * float(issued_size)
                # Convert to ₹ Crores (same convention as yfinance branch)
                market_cap = round(market_cap_raw / 10000000, 2)
        except Exception:
            market_cap = None

        return {
            "price": float(last_price),
            "pchange": float(pchange) if pchange is not None else None,
            "volume": float(volume) if volume is not None else None,
            "sector": sector,
            "market_cap": market_cap,
        }
    except Exception as e:
        print(f"   ⚠️ NSE quote failed for {symbol}: {e}")
        return None


def get_market_intelligence_from_nse(symbol, cap_size):
    """Fallback: build intelligence snapshot using NSE quote data.

    This is used when yfinance cannot provide a usable price history.
    """

    nse = fetch_price_from_nse(symbol)
    if not nse:
        return None

    latest_price = nse["price"]
    pchange = nse["pchange"]
    market_cap = nse["market_cap"]
    sector = nse["sector"]

    # Use daily percentage change as a short-term momentum proxy
    if pchange is not None:
        momentum = pchange / 100.0
    else:
        momentum = 0.0

    # Without a long history we treat volume as neutral (1.0x)
    vol_surge = 1.0

    if pchange is not None:
        simple_sentiment = 1.0 if momentum > 0.02 else (-1.0 if momentum < -0.02 else 0.0)
    else:
        simple_sentiment = 0.0

    # Fetch news (same logic as primary path)
    news_text = None
    news_titles = []
    news_sentiment = None
    news_types = []

    try:
        if HAS_COMPREHENSIVE_NEWS:
            news_text, news_titles = fetch_news_comprehensive(symbol)
        else:
            news_text, news_titles = fetch_news(symbol)

        if news_text:
            news_sentiment = analyze_news_sentiment(news_text)
            news_types = classify_news_type(news_text)
    except Exception as e:
        print(f"   ⚠️  Failed to fetch news (NSE fallback): {str(e)}")

    print(
        f"📊 {symbol}: (NSE) Price=₹{latest_price:.2f}, "
        f"Momentum≈{momentum*100:.1f}% (1D), Volume≈{vol_surge:.2f}x, "
        f"Trend={'📈' if simple_sentiment > 0 else '📉' if simple_sentiment < 0 else '➡️'}"
    )

    return {
        "ticker": symbol,
        "cap": cap_size,
        "sector": sector,
        "price": latest_price,
        "market_cap": market_cap,
        "mom": momentum,
        "vol": vol_surge,
        "sent": simple_sentiment,
        "news": news_text,
        "news_sentiment": news_sentiment,
        "news_types": news_types,
        "has_data": True,
    }


def get_market_intelligence(symbol, cap_size):
    try:
        # Use current ticker (handles renamed companies)
        current_symbol = get_current_ticker(symbol)

        # Check if delisted or pump & dump
        if is_delisted(current_symbol):
            logger.debug(f"{symbol} is delisted, skipping")
            return None

        stock = yf.Ticker(f"{current_symbol}.NS")
        df = stock.history(period="7mo", auto_adjust=True)

        # Check if we have sufficient data
        has_data = not df.empty and len(df) >= 20

        if has_data:
            # Normal case: sufficient data available
            latest_price = df['Close'].iloc[-1]
            momentum = (df['Close'].iloc[-1] - df['Close'].iloc[0]) / df['Close'].iloc[0]
            avg_vol = df['Volume'].iloc[:-1].tail(20).mean()
            vol_surge = df['Volume'].iloc[-1] / avg_vol if avg_vol > 0 else 1.0
            recent_change = (df['Close'].iloc[-1] - df['Close'].iloc[-6]) / df['Close'].iloc[-6]
            simple_sentiment = 1.0 if recent_change > 0.02 else (-1.0 if recent_change < -0.02 else 0.0)

            # Get market capitalization and sector
            market_cap = None
            sector = "Unknown"
            try:
                info = stock.info
                market_cap_raw = info.get('marketCap', None)
                if market_cap_raw:
                    market_cap = round(market_cap_raw / 10000000, 2)

                # Get sector information
                sector = info.get('sector', info.get('industry', 'Unknown'))
            except Exception:
                pass

            # Fetch news
            news_text = None
            news_titles = []
            news_sentiment = None
            news_types = []

            try:
                if HAS_COMPREHENSIVE_NEWS:
                    # Use comprehensive news sources (70+ sources)
                    news_text, news_titles = fetch_news_comprehensive(symbol)
                else:
                    # Use basic news (Yahoo + Google)
                    news_text, news_titles = fetch_news(symbol)

                if news_text:
                    news_sentiment = analyze_news_sentiment(news_text)
                    news_types = classify_news_type(news_text)
            except Exception as e:
                print(f"   ⚠️  Failed to fetch news: {str(e)}")

            print(f"📊 {symbol}: Price=₹{latest_price:.2f}, Momentum={momentum*100:.1f}%, Volume={vol_surge:.2f}x, Trend={'📈' if simple_sentiment > 0 else '📉' if simple_sentiment < 0 else '➡️'}")

            return {
                "ticker": symbol,
                "cap": cap_size,
                "sector": sector,
                "price": latest_price,
                "market_cap": market_cap,
                "mom": momentum,
                "vol": vol_surge,
                "sent": simple_sentiment,
                "news": news_text,
                "news_sentiment": news_sentiment,
                "news_types": news_types,
                "has_data": True
            }
        else:
            # No data from yfinance – try NSE fallback
            print(f"⚠️  {symbol}: No data from yfinance, trying NSE fallback...")
            alt = get_market_intelligence_from_nse(symbol, cap_size)
            if alt:
                return alt

            # Still nothing: add with NA/default values
            print(f"⚠️  {symbol}: No data from any source - adding with NA values")
            return {
                "ticker": symbol,
                "cap": cap_size,
                "sector": "Unknown",
                "price": None,
                "market_cap": None,
                "mom": 0.0,
                "vol": 0.0,
                "sent": 0.0,
                "news": None,
                "news_sentiment": None,
                "news_types": [],
                "has_data": False
            }
    except Exception as e:
        # yfinance raised an exception – attempt NSE-based fallback
        print(f"⚠️  {symbol}: yfinance error '{e}'. Trying NSE fallback...")
        alt = get_market_intelligence_from_nse(symbol, cap_size)
        if alt:
            return alt

        print(f"⚠️  {symbol}: Error - adding with NA values")
        return {
            "ticker": symbol,
            "cap": cap_size,
            "sector": "Unknown",
            "price": None,
            "market_cap": None,
            "mom": 0.0,
            "vol": 0.0,
            "sent": 0.0,
            "news": None,
            "news_sentiment": None,
            "news_types": [],
            "has_data": False
        }

def calculate_signal_score(data):
    """Calculate signal and score for a stock"""
    if not data.get('has_data', True):
        return "❄️ N/A", 0

    signal = "❄️ Neutral"
    if data['mom'] > 0.10 and data['vol'] > 1.2:
        signal = "🚀 Strong Buy"
    elif data['mom'] > 0.05 or data['vol'] > 1.1:
        signal = "👀 Watch"

    score = 0
    if signal == "🚀 Strong Buy":
        score += 1000
    elif signal == "👀 Watch":
        score += 500
    score += data['mom'] * 500
    score += data['vol'] * 50
    score = round(score, 2)

    return signal, score


def get_institutional_holdings(symbol):
    """Fetch FII/DII/Promoter holdings and trends from multiple public sources.

    Current implementation uses Screener's shareholding pattern as the primary
    source because it is indexed by NSE symbol and provides a clean history of
    FII/ DII / Promoter percentages by quarter. The design is extensible so that
    additional sources (Moneycontrol, NSE, BSE, etc.) can be layered in later
    and reconciled.

    Returns a dict with:
        fii_pct, dii_pct, promoter_pct, mf_pct,
        fii_trend, dii_trend, inst_confidence.
    """

    base = {
        "fii_pct": None,
        "dii_pct": None,
        "promoter_pct": None,
        "mf_pct": None,
        "fii_trend": "N/A",
        "dii_trend": "N/A",
        "inst_confidence": "N/A",
    }

    try:
        # Primary source: Screener shareholding pattern (quarterly)
        url = f"https://www.screener.in/company/{symbol}/consolidated/"
        headers = {
            "User-Agent": USER_AGENT,
            "Referer": "https://www.screener.in/",
        }
        resp = SESSION.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            print(f"   ⚠️  Screener shareholding fetch failed for {symbol}: HTTP {resp.status_code}")
            return base

        html = resp.text
        if "Shareholding Pattern" not in html:
            # No structured shareholding section
            return base

        # Narrow down to the shareholding pattern block
        start_idx = html.find("Shareholding Pattern")
        block = html[start_idx:]

        # Focus on the dedicated quarterly shareholding div. The structure is:
        # <div id="quarterly-shp"> ... </div>
        # <div id="yearly-shp"> ... </div>
        q_div = '<div id="quarterly-shp">'
        y_div = '<div id="yearly-shp">'

        q_start = block.find(q_div)
        if q_start == -1:
            return base

        y_start = block.find(y_div, q_start)
        if y_start != -1:
            q_block = block[q_start:y_start]
        else:
            q_block = block[q_start:]

        def _extract_series(label):
            """Extract all percentage values for a given row label from the quarterly block.

            We anchor on the row label (e.g. "FIIs", "DIIs", "Promoters") and capture
            everything until the end of that table row (</tr>). This avoids accidentally
            pulling in values from other rows or the yearly table.
            """
            try:
                m = re.search(rf"{label}.*?</tr>", q_block, re.S)
                if not m:
                    return []
                # Extract all occurrences of "xx.xx%" within that row
                nums = re.findall(r"(\d+\.\d+)%", m.group(0))
                return [float(x) for x in nums]
            except Exception:
                return []

        fii_series = _extract_series("FIIs")
        dii_series = _extract_series("DIIs")
        prom_series = _extract_series("Promoters")
        # Some companies expose a dedicated Mutual Funds row; where present,
        # we treat that as the most accurate MF holding %.
        mf_series = _extract_series("Mutual Funds") or _extract_series("Mutual Fund")

        def _latest_and_prev(series):
            if not series:
                return None, None
            if len(series) == 1:
                return series[-1], None
            return series[-1], series[-2]

        fii_cur, fii_prev = _latest_and_prev(fii_series)
        dii_cur, dii_prev = _latest_and_prev(dii_series)
        prom_cur, prom_prev = _latest_and_prev(prom_series)
        mf_cur, mf_prev = _latest_and_prev(mf_series)

        def _trend(cur, prev):
            if cur is None or prev is None:
                return "N/A"
            delta = cur - prev
            # Use a small threshold to avoid classifying noise as a real trend
            if delta > 0.3:
                return "Increasing"
            if delta < -0.3:
                return "Decreasing"
            return "Stable"

        fii_trend = _trend(fii_cur, fii_prev)
        dii_trend = _trend(dii_cur, dii_prev)

        inst_total = None
        if fii_cur is not None and dii_cur is not None:
            inst_total = fii_cur + dii_cur

        # Institutional confidence heuristic:
        # - HIGH   : strong institutional ownership (>60%) and at least one trend increasing
        # - MEDIUM : moderate ownership (35–60%) or mixed trends
        # - LOW    : low ownership (<35%) or both trends decreasing
        confidence = "N/A"
        if inst_total is not None:
            if inst_total >= 60:
                if fii_trend == "Increasing" or dii_trend == "Increasing":
                    confidence = "High"
                elif fii_trend == "Decreasing" and dii_trend == "Decreasing":
                    confidence = "Medium"
                else:
                    confidence = "High"
            elif inst_total >= 35:
                if fii_trend == "Decreasing" and dii_trend == "Decreasing":
                    confidence = "Low"
                else:
                    confidence = "Medium"
            else:
                confidence = "Low"

        result = base.copy()
        result.update(
            {
                "fii_pct": fii_cur,
                "dii_pct": dii_cur,
                "promoter_pct": prom_cur,
                # MF % (if Screener exposes a Mutual Funds row)
                "mf_pct": mf_cur,
                "fii_trend": fii_trend,
                "dii_trend": dii_trend,
                "inst_confidence": confidence,
            }
        )
        return result
    except Exception as e:
        print(f"   ⚠️  Failed to fetch institutional holdings for {symbol}: {e}")
        return base


def calculate_institutional_score(data):
    """Calculate a 0–100 institutional quality score.

    Combines:
      - Absolute FII + DII ownership (base institutional presence)
      - Promoter holding quality
      - Mutual fund holding (when explicitly available)
      - FII / DII trends
    """

    fii = data.get("fii_pct") or 0.0
    dii = data.get("dii_pct") or 0.0
    prom = data.get("promoter_pct") or 0.0
    mf = data.get("mf_pct")
    fii_trend = (data.get("fii_trend") or "N/A").lower()
    dii_trend = (data.get("dii_trend") or "N/A").lower()

    inst_total = fii + dii
    score = 0.0

    # 1) Base institutional ownership (FII + DII) → up to ~60 points
    if inst_total >= 60:
        score += 60
    elif inst_total >= 35:
        # 35–60% → 40–60 points (linear)
        score += 40 + (inst_total - 35) * (20.0 / 25.0)
    elif inst_total >= 10:
        # 10–35% → 0–40 points (linear)
        score += (inst_total - 10) * (40.0 / 25.0)

    # 2) Promoter holding quality → up to ~15 points
    #    Sweet spot ~30–60%: enough skin in the game, not overly concentrated.
    if 30 <= prom <= 60:
        score += 15
    elif 10 <= prom < 30 or 60 < prom <= 75:
        score += 8
    elif prom > 75:
        score += 3

    # 3) Mutual fund holdings (if explicitly reported) → up to 10 points
    if mf is not None:
        if mf >= 15:
            score += 10
        elif mf >= 8:
            score += 7
        elif mf >= 3:
            score += 4
        elif mf >= 1:
            score += 2

    # 4) Flows / trend effects → up to +/- 15 points
    if "increasing" in fii_trend:
        score += 6
    elif "decreasing" in fii_trend:
        score -= 4

    if "increasing" in dii_trend:
        score += 6
    elif "decreasing" in dii_trend:
        score -= 4

    # Clamp and round
    score = max(0.0, min(100.0, score))
    return round(score, 1)


def create_excel_report(stocks_data, filename):
    """Create formatted Excel report with all stock data"""
    print("\n" + "="*70)
    print("📊 CREATING EXCEL REPORT")
    print("="*70)

    # Prepare data for DataFrame
    excel_data = []

    for stock in stocks_data:
        signal, score = calculate_signal_score(stock)

        row = {
            'Rank': stock.get('rank', 0),
            'Symbol': stock['ticker'],
            'Price (₹)': round(stock['price'], 2) if stock.get('price') else 0,
            'Momentum (%)': round(stock['mom'] * 100, 2),
            'Volume Surge': round(stock['vol'], 2),
            'Trend': '📈' if stock['sent'] > 0 else ('📉' if stock['sent'] < 0 else '➡️'),
            'Score': score,
            'Signal': signal,
            'Sector': stock.get('sector', 'Unknown'),
            'Market Cap': stock.get('cap', 'N/A'),
            'Capital Market (₹Cr)': round(stock['market_cap'], 2) if stock.get('market_cap') else 0,
            'News Sentiment': stock.get('news_sentiment', 'N/A'),
            'News Types': ', '.join(stock.get('news_types', [])) if stock.get('news_types') else 'N/A',
            'News & Updates': stock.get('news', 'N/A')[:500] if stock.get('news') else 'N/A',
            'Consensus': stock.get('consensus', 'N/A'),
            'Ratings': stock.get('ratings', 'N/A'),
            'FII Holdings %': stock.get('fii_pct'),
            'DII Holdings %': stock.get('dii_pct'),
            'Promoter Holdings %': stock.get('promoter_pct'),
            'Mutual Fund Holdings %': stock.get('mf_pct'),
            'FII Trend': stock.get('fii_trend', 'N/A'),
            'DII Trend': stock.get('dii_trend', 'N/A'),
            'Institutional Confidence': stock.get('inst_confidence', 'N/A'),
            'Institutional Score': stock.get('institutional_score'),
            'Last Updated': datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        excel_data.append(row)

    # Create DataFrame
    df = pd.DataFrame(excel_data)

    # Sort by Rank
    df = df.sort_values('Rank', ascending=True).reset_index(drop=True)

    # Create Excel file with formatting
    print(f"\n💾 Creating Excel file: {filename}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Set column widths
    column_widths = {
        'A': 8,   # Rank
        'B': 15,  # Symbol
        'C': 12,  # Price
        'D': 15,  # Momentum
        'E': 15,  # Volume Surge
        'F': 10,  # Trend
        'G': 10,  # Score
        'H': 18,  # Signal
        'I': 25,  # Sector
        'J': 15,  # Market Cap
        'K': 18,  # Capital Market
        'L': 18,  # News Sentiment
        'M': 25,  # News Types
        'N': 50,  # News & Updates
        'O': 18,  # Consensus
        'P': 18,  # Ratings
        'Q': 16,  # FII Holdings %
        'R': 16,  # DII Holdings %
        'S': 20,  # Promoter Holdings %
        'T': 20,  # Mutual Fund Holdings %
        'U': 20,  # FII Trend
        'V': 20,  # DII Trend
        'W': 24,  # Institutional Confidence
        'X': 16,  # Institutional Score
        'Y': 20,  # Last Updated
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save file
    wb.save(filename)

    print("✅ Excel file created successfully!")
    print(f"📁 Location: {os.path.abspath(filename)}")
    print(f"📊 Total stocks: {len(df)}")

    # Show top 5 stocks
    print("\n📈 Top 5 ranked stocks:")
    for idx, row in df.head(5).iterrows():
        print(f"   {int(row['Rank'])}. {row['Symbol']:15s} - ₹{row['Price (₹)']:.2f} (+{row['Momentum (%)']:.1f}%) - {row['Sector']}")

    print("\n" + "="*70)
    print("✅ EXCEL REPORT CREATION COMPLETE!")
    print("="*70)

    return filename

if __name__ == "__main__":
    import time
    start_time = time.time()

    print("\n" + "="*70)
    print("📈 MARKET INTELLIGENCE BOT - EXCEL VERSION")
    print("="*70)
    print("🏆 Intelligent Multi-Factor Ranking: ENABLED")
    print("📊 Output Format: Excel (.xlsx)")
    if HAS_RANKING_ENGINE:
        print("✅ Using optimized ranking engine")
    else:
        print("⚠️  Ranking engine not available - using serial ranking")
    print("="*70 + "\n")

    # Create output directory if it doesn't exist
    excel_output_dir = os.path.join(os.path.dirname(__file__), "excel")
    os.makedirs(excel_output_dir, exist_ok=True)
    print(f"📁 Excel output directory: {excel_output_dir}\n")

    # Load all NSE stocks (Nifty 100, Midcap 150, Smallcap 250)
    if get_all_stocks_with_classification is None:
        print("❌ ERROR: Stock data module not available!")
        print("Please ensure data/nse_stocks_650.py exists and is accessible.")
        exit(1)

    watchlist = get_all_stocks_with_classification()

    print(f"📊 Total stocks to analyze: {len(watchlist)}")
    print(f"⏱️  Estimated time: ~{len(watchlist) * 2 // 60} minutes\n")

    # Statistics tracking
    stats = {
        "total": len(watchlist),
        "success": 0,
        "errors": 0,
        "strong_buy": 0,
        "watch": 0,
        "neutral": 0
    }

    # PHASE 1: Data Collection
    print("="*70)
    print("📥 PHASE 1: DATA COLLECTION")
    print("="*70)

    all_stocks_data = []
    for idx, (symbol, cap_size) in enumerate(watchlist, 1):
        print(f"\n[{idx}/{stats['total']}] Analyzing {symbol} ({cap_size})...")
        data = get_market_intelligence(symbol, cap_size)

        # Fetch analyst ratings if available
        if HAS_ANALYST_RATINGS and data.get('has_data', True):
            try:
                print(f"   📊 Fetching analyst ratings for {symbol}...")
                ratings_data = aggregate_all_analyst_ratings(symbol)
                if ratings_data['has_data']:
                    data['consensus'] = ratings_data['consensus']
                    data['ratings'] = f"{ratings_data['rating_numeric']:.2f}/5.0 ({ratings_data['analyst_count']} analysts)"
                    print(f"   ✅ Ratings: {ratings_data['consensus']} | {data['ratings']}")
                else:
                    data['consensus'] = "No Consensus"
                    data['ratings'] = "N/A"
            except Exception as e:
                print(f"   ⚠️  Failed to fetch ratings: {str(e)}")
                data['consensus'] = "N/A"
                data['ratings'] = "N/A"
        else:
            data['consensus'] = "N/A"
            data['ratings'] = "N/A"

        # Fetch institutional holdings (FII/DII/Promoters/MFs + trends) and compute score
        try:
            print(f"   📊 Fetching institutional holdings for {symbol}...")
            holdings = get_institutional_holdings(symbol)
            data.update(holdings)
            data["institutional_score"] = calculate_institutional_score(data)

            if holdings.get("fii_pct") is not None and holdings.get("dii_pct") is not None:
                prom_disp = holdings.get("promoter_pct")
                prom_disp = prom_disp if prom_disp is not None else 0.0
                print(
                    f"   ✅ Holdings: FII {holdings['fii_pct']:.2f}%, "
                    f"DII {holdings['dii_pct']:.2f}%, Promoter {prom_disp:.2f}% | "
                    f"InstScore: {data['institutional_score']}"
                )
        except Exception as e:
            print(f"   ⚠️  Failed to fetch institutional holdings: {str(e)}")

        all_stocks_data.append(data)

        if data.get('has_data'):
            stats["success"] += 1
        else:
            stats["errors"] += 1

        time.sleep(0.3)  # Rate limiting

    # PHASE 2: Intelligent Ranking
    print("\n" + "="*70)
    print("🏆 PHASE 2: INTELLIGENT MULTI-FACTOR RANKING")
    print("="*70)

    if HAS_RANKING_ENGINE:
        # Use optimized ranking engine
        print("⚡ Using optimized ranking engine...")
        ranked_stocks = rank_stocks(all_stocks_data)
        print(f"✅ Ranked {len(ranked_stocks)} stocks using intelligent multi-factor algorithm")
    else:
        # Fallback: simple ranking by score
        print("⚠️  Using fallback ranking (by score)...")
        for stock in all_stocks_data:
            signal, score = calculate_signal_score(stock)
            stock['signal'] = signal
            stock['score'] = score

        ranked_stocks = sorted(all_stocks_data, key=lambda x: x.get('score', 0), reverse=True)
        for idx, stock in enumerate(ranked_stocks, 1):
            stock['rank'] = idx
        print(f"✅ Ranked {len(ranked_stocks)} stocks by score")

    # Count signals
    for stock in ranked_stocks:
        signal = stock.get('signal', '❄️ Neutral')
        if "Strong Buy" in signal:
            stats["strong_buy"] += 1
        elif "Watch" in signal:
            stats["watch"] += 1
        else:
            stats["neutral"] += 1

    # PHASE 3: Create Excel Report
    print("\n" + "="*70)
    print("📊 PHASE 3: CREATING EXCEL REPORT")
    print("="*70)

    # Generate timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = os.path.join(excel_output_dir, f"market_analysis_{timestamp}.xlsx")

    create_excel_report(ranked_stocks, excel_filename)

    # Final Summary
    elapsed_time = time.time() - start_time
    print("\n" + "="*70)
    print("📊 FINAL SUMMARY")
    print("="*70)
    print(f"⏱️  Total Time: {elapsed_time/60:.1f} minutes")
    print(f"📈 Total Stocks: {stats['total']}")
    print(f"✅ Successful: {stats['success']}")
    print(f"❌ Errors: {stats['errors']}")
    print("\n🎯 Signal Distribution:")
    print(f"   🚀 Strong Buy: {stats['strong_buy']}")
    print(f"   👀 Watch: {stats['watch']}")
    print(f"   ❄️  Neutral/N/A: {stats['neutral']}")
    print(f"\n📁 Excel Report: {excel_filename}")
    print("="*70)
    print("✅ MARKET ANALYSIS COMPLETE!")
    print("="*70)
